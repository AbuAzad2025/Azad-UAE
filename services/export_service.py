"""
Export Service - خدمة التصدير
تصدير التقارير إلى PDF, Excel, CSV
"""

from datetime import datetime, timezone
from io import BytesIO, StringIO
import csv
import logging

logger = logging.getLogger(__name__)


class ExportService:
    """خدمة تصدير التقارير"""

    @staticmethod
    def export_to_csv(data, headers, filename="export.csv"):
        """
        تصدير البيانات إلى CSV

        Args:
            data (list): قائمة البيانات
            headers (list): عناوين الأعمدة
            filename (str): اسم الملف

        Returns:
            BytesIO: ملف CSV
        """
        # Create a string buffer for CSV writing
        str_output = StringIO()
        writer = csv.writer(str_output)
        writer.writerow(headers)

        for row in data:
            writer.writerow(row)

        csv_content = str_output.getvalue()

        # Create bytes buffer with BOM for Excel compatibility
        output = BytesIO()
        output.write(b"\xef\xbb\xbf")
        output.write(csv_content.encode("utf-8"))

        output.seek(0)
        return output

    @staticmethod
    def export_to_xlsx(data, headers, filename="export.xlsx", sheet_name="Report"):
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = (sheet_name or "Report")[:31]

        ws.append(list(headers))
        for row in data:
            ws.append(list(row))

        for col_idx, header in enumerate(headers, start=1):
            max_len = len(str(header or ""))
            for cell in ws[get_column_letter(col_idx)]:
                try:
                    v = "" if cell.value is None else str(cell.value)
                    if len(v) > max_len:
                        max_len = len(v)
                except Exception:
                    continue
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 60)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def export_purchases_to_csv(purchases):
        """تصدير المشتريات إلى CSV"""
        headers = [
            "الرقم",
            "الباقة",
            "العميل",
            "البريد الإلكتروني",
            "المبلغ",
            "طريقة الدفع",
            "الحالة",
            "التاريخ",
        ]

        data = []
        for purchase in purchases:
            data.append(
                [
                    purchase.id,
                    purchase.package.name_ar if purchase.package else "N/A",
                    purchase.customer_name,
                    purchase.customer_email,
                    f"${purchase.amount_paid}",
                    purchase.payment_method,
                    purchase.payment_status,
                    (purchase.created_at.strftime("%Y-%m-%d %H:%M") if purchase.created_at else "N/A"),
                ]
            )

        return ExportService.export_to_csv(data, headers, "purchases.csv")

    @staticmethod
    def export_donations_to_csv(donations):
        """تصدير التبرعات إلى CSV"""
        headers = [
            "الرقم",
            "المتبرع",
            "البريد الإلكتروني",
            "المبلغ",
            "طريقة الدفع",
            "الحالة",
            "التاريخ",
        ]

        data = []
        for donation in donations:
            data.append(
                [
                    donation.id,
                    donation.donor_name or "مجهول",
                    donation.donor_email or "N/A",
                    f"${donation.amount_usd}",
                    donation.payment_method,
                    donation.status,
                    (donation.created_at.strftime("%Y-%m-%d %H:%M") if donation.created_at else "N/A"),
                ]
            )

        return ExportService.export_to_csv(data, headers, "donations.csv")

    @staticmethod
    def export_cards_to_csv(cards):
        """تصدير البطاقات إلى CSV (معلومات عامة فقط)"""
        headers = [
            "الرقم",
            "العميل",
            "البريد الإلكتروني",
            "البطاقة",
            "النوع",
            "المبلغ",
            "الحالة",
            "التاريخ",
        ]

        data = []
        for card in cards:
            data.append(
                [
                    card.id,
                    card.customer_name,
                    card.customer_email,
                    (card.get_card_display() if hasattr(card, "get_card_display") else "N/A"),
                    card.card_type or "Unknown",
                    f"${card.amount}" if card.amount else "N/A",
                    card.status,
                    (card.created_at.strftime("%Y-%m-%d %H:%M") if card.created_at else "N/A"),
                ]
            )

        return ExportService.export_to_csv(data, headers, "cards.csv")

    @staticmethod
    def generate_pdf_report(title, data, filename="report.pdf"):
        """
        إنشاء تقرير PDF (تنفيذ بسيط - يمكن تحسينه باستخدام ReportLab)

        Args:
            title (str): عنوان التقرير
            data (dict): بيانات التقرير
            filename (str): اسم الملف

        Returns:
            str: HTML content (للطباعة أو التحويل لـ PDF)
        """
        html = f"""
        <!DOCTYPE html>
        <html dir="rtl" lang="ar">
        <head>
            <meta charset="UTF-8">
            <title>{title}</title>
            <style>
                body {{
                    font-family: 'Arial', sans-serif;
                    direction: rtl;
                    padding: 20px;
                }}
                h1 {{
                    color: #667eea;
                    border-bottom: 3px solid #667eea;
                    padding-bottom: 10px;
                }}
                .stats {{
                    display: grid;
                    grid-template-columns: repeat(3, 1fr);
                    gap: 20px;
                    margin: 30px 0;
                }}
                .stat-box {{
                    background: #f8f9fa;
                    padding: 20px;
                    border-radius: 10px;
                    text-align: center;
                    border: 2px solid #667eea;
                }}
                .stat-number {{
                    font-size: 2rem;
                    color: #667eea;
                    font-weight: bold;
                }}
                .stat-label {{
                    color: #666;
                    margin-top: 10px;
                }}
                table {{
                    width: 100%;
                    border-collapse: collapse;
                    margin-top: 20px;
                }}
                th, td {{
                    border: 1px solid #ddd;
                    padding: 12px;
                    text-align: right;
                }}
                th {{
                    background-color: #667eea;
                    color: white;
                }}
                tr:nth-child(even) {{
                    background-color: #f8f9fa;
                }}
                .footer {{
                    margin-top: 40px;
                    text-align: center;
                    color: #999;
                    font-size: 0.9rem;
                }}
            </style>
        </head>
        <body>
            <h1>{title}</h1>
            <p><strong>تاريخ التقرير:</strong> {datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M")}</p>
            
            <div class="stats">
                {ExportService._generate_stats_html(data.get("stats", {}))}
            </div>
            
            {ExportService._generate_table_html(data.get("table_data", []), data.get("table_headers", []))}
            
            <div class="footer">
                <p>تم إنشاؤه بواسطة Azad Systems - الخزينة السرية</p>
            </div>
        </body>
        </html>
        """

        return html

    @staticmethod
    def _generate_stats_html(stats):
        """إنشاء HTML للإحصائيات"""
        html = ""
        for key, value in stats.items():
            html += f"""
            <div class="stat-box">
                <div class="stat-number">{value}</div>
                <div class="stat-label">{key}</div>
            </div>
            """
        return html

    @staticmethod
    def _generate_table_html(data, headers):
        """إنشاء HTML للجدول"""
        if not data or not headers:
            return ""

        html = "<table><thead><tr>"
        for header in headers:
            html += f"<th>{header}</th>"
        html += "</tr></thead><tbody>"

        for row in data:
            html += "<tr>"
            for cell in row:
                html += f"<td>{cell}</td>"
            html += "</tr>"

        html += "</tbody></table>"
        return html
